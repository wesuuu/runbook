"""Data export service with strategy pattern for format extensibility.

Transforms run graph + execution_data into flat tabular rows,
then serializes via pluggable format strategies (CSV, Excel, JSON).
"""

import csv
import io
import json
from abc import ABC, abstractmethod
from typing import Any


# ── Column definitions ──

LONG_COLUMNS: list[dict[str, str]] = [
    {"key": "run_name", "label": "Run Name", "group": "metadata"},
    {"key": "protocol_name", "label": "Protocol", "group": "metadata"},
    {"key": "run_status", "label": "Run Status", "group": "metadata"},
    {"key": "step_number", "label": "Step #", "group": "step"},
    {"key": "step_id", "label": "Step ID", "group": "step"},
    {"key": "step_name", "label": "Step Name", "group": "step"},
    {"key": "category", "label": "Category", "group": "step"},
    {"key": "role", "label": "Role", "group": "step"},
    {"key": "description", "label": "Description", "group": "step"},
    {"key": "duration_min", "label": "Duration (min)", "group": "step"},
    {"key": "param_name", "label": "Parameter", "group": "data"},
    {"key": "param_value", "label": "Value", "group": "data"},
    {"key": "completed_by", "label": "Completed By", "group": "audit"},
    {"key": "completed_at", "label": "Completed At", "group": "audit"},
    {"key": "edited", "label": "Edited", "group": "audit"},
    {"key": "original_value", "label": "Original Value", "group": "audit"},
    {"key": "edited_by", "label": "Edited By", "group": "audit"},
    {"key": "edited_at", "label": "Edited At", "group": "audit"},
]

WIDE_BASE_COLUMNS: list[dict[str, str]] = [
    {"key": "run_name", "label": "Run Name", "group": "metadata"},
    {"key": "protocol_name", "label": "Protocol", "group": "metadata"},
    {"key": "run_status", "label": "Run Status", "group": "metadata"},
    {"key": "step_number", "label": "Step #", "group": "step"},
    {"key": "step_id", "label": "Step ID", "group": "step"},
    {"key": "step_name", "label": "Step Name", "group": "step"},
    {"key": "category", "label": "Category", "group": "step"},
    {"key": "role", "label": "Role", "group": "step"},
    {"key": "description", "label": "Description", "group": "step"},
    {"key": "duration_min", "label": "Duration (min)", "group": "step"},
]

WIDE_AUDIT_COLUMNS: list[dict[str, str]] = [
    {"key": "completed_by", "label": "Completed By", "group": "audit"},
    {"key": "completed_at", "label": "Completed At", "group": "audit"},
    {"key": "edited", "label": "Edited", "group": "audit"},
    {"key": "edited_by", "label": "Edited By", "group": "audit"},
    {"key": "edited_at", "label": "Edited At", "group": "audit"},
]


# ── Export strategies ──


class ExportStrategy(ABC):
    """Base class for export format strategies."""

    @property
    @abstractmethod
    def media_type(self) -> str: ...

    @property
    @abstractmethod
    def file_extension(self) -> str: ...

    @abstractmethod
    def export(
        self,
        columns: list[dict[str, str]],
        rows: list[dict[str, Any]],
        metadata: dict[str, Any],
    ) -> bytes: ...


class CsvExportStrategy(ExportStrategy):
    @property
    def media_type(self) -> str:
        return "text/csv"

    @property
    def file_extension(self) -> str:
        return "csv"

    def export(self, columns, rows, metadata) -> bytes:
        output = io.StringIO()
        col_keys = [c["key"] for c in columns]
        col_labels = [c["label"] for c in columns]
        writer = csv.writer(output)
        writer.writerow(col_labels)
        for row in rows:
            writer.writerow([row.get(k, "") for k in col_keys])
        return output.getvalue().encode("utf-8")


class ExcelExportStrategy(ExportStrategy):
    @property
    def media_type(self) -> str:
        return (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def export(self, columns, rows, metadata) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()

        # — Data sheet —
        ws = wb.active
        ws.title = "Run Data"

        col_keys = [c["key"] for c in columns]
        col_labels = [c["label"] for c in columns]

        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=10)

        for ci, label in enumerate(col_labels, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(col_keys, 1):
                ws.cell(row=ri, column=ci, value=row.get(key, ""))

        # Auto-fit column widths
        for ci, key in enumerate(col_keys, 1):
            max_len = len(col_labels[ci - 1])
            for row in rows[:200]:  # sample first 200 rows
                val_len = len(str(row.get(key, "")))
                if val_len > max_len:
                    max_len = val_len
            letter = ws.cell(row=1, column=ci).column_letter
            ws.column_dimensions[letter].width = min(max_len + 3, 50)

        # — Metadata sheet —
        ws_meta = wb.create_sheet("Metadata")
        meta_rows: list[tuple[str, str]] = [
            ("Export Date", str(metadata.get("export_date", ""))),
            ("Run Count", str(metadata.get("run_count", 0))),
            ("Layout", str(metadata.get("layout", ""))),
        ]
        for run_info in metadata.get("runs", []):
            meta_rows.append(("", ""))
            meta_rows.append(("Run Name", run_info.get("name", "")))
            meta_rows.append(("Run Status", run_info.get("status", "")))
            meta_rows.append(("Protocol", run_info.get("protocol_name", "")))
            meta_rows.append(("Created", str(run_info.get("created_at", ""))))
            meta_rows.append(("Updated", str(run_info.get("updated_at", ""))))

        for ri, (k, v) in enumerate(meta_rows, 1):
            ws_meta.cell(row=ri, column=1, value=k)
            ws_meta.cell(row=ri, column=2, value=v)
            if k:
                ws_meta.cell(row=ri, column=1).font = Font(bold=True)

        ws_meta.column_dimensions["A"].width = 20
        ws_meta.column_dimensions["B"].width = 40

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class JsonExportStrategy(ExportStrategy):
    @property
    def media_type(self) -> str:
        return "application/json"

    @property
    def file_extension(self) -> str:
        return "json"

    def export(self, columns, rows, metadata) -> bytes:
        data = {
            "metadata": metadata,
            "columns": columns,
            "rows": rows,
        }
        return json.dumps(data, indent=2, default=str).encode("utf-8")


# Strategy registry — add new formats here
EXPORT_STRATEGIES: dict[str, ExportStrategy] = {
    "csv": CsvExportStrategy(),
    "xlsx": ExcelExportStrategy(),
    "json": JsonExportStrategy(),
}


def get_strategy(format_key: str) -> ExportStrategy:
    strategy = EXPORT_STRATEGIES.get(format_key)
    if not strategy:
        raise ValueError(f"Unknown export format: {format_key}")
    return strategy


# ── Data transformation ──


def _extract_ordered_steps(graph: dict) -> list[dict]:
    """Extract unit op steps from graph in x-position order."""
    nodes = graph.get("nodes", [])
    unit_ops = sorted(
        [n for n in nodes if n.get("type") == "unitOp"],
        key=lambda n: n.get("position", {}).get("x", 0),
    )
    swim_lanes = {
        n["id"]: n for n in nodes if n.get("type") == "swimLane"
    }

    steps = []
    for i, node in enumerate(unit_ops, 1):
        data = node.get("data", {})
        parent_id = node.get("parentId")
        role = ""
        if parent_id and parent_id in swim_lanes:
            role = swim_lanes[parent_id].get("data", {}).get("label", "")

        steps.append({
            "step_number": i,
            "step_id": node["id"],
            "step_name": data.get("label", "Unnamed"),
            "category": data.get("category", ""),
            "role": role,
            "description": data.get("description", ""),
            "duration_min": data.get("duration_min"),
        })
    return steps


def build_export_data(
    runs: list[dict[str, Any]],
    layout: str,
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """Build export columns and rows from run data.

    Args:
        runs: List of run dicts with keys: name, status, graph,
              execution_data, user_map, protocol_name
        layout: "long" or "wide"

    Returns:
        (columns, rows) tuple.
    """
    if layout == "long":
        return _build_long_format(runs)
    return _build_wide_format(runs)


def _build_step_base(
    run: dict[str, Any],
    step: dict[str, Any],
    step_exec: dict[str, Any],
    user_map: dict[str, str],
) -> dict[str, Any]:
    """Build the common base row dict for a step."""
    return {
        "run_name": run["name"],
        "protocol_name": run.get("protocol_name", ""),
        "run_status": run.get("status", ""),
        "step_number": step["step_number"],
        "step_id": step["step_id"],
        "step_name": step["step_name"],
        "category": step["category"],
        "role": step["role"],
        "description": step["description"],
        "duration_min": step["duration_min"] or "",
        "completed_by": user_map.get(
            step_exec.get("completed_by_user_id", ""), ""
        ),
        "completed_at": step_exec.get("completed_at", ""),
    }


def _build_long_format(
    runs: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """One row per step x parameter."""
    columns = list(LONG_COLUMNS)
    rows: list[dict[str, Any]] = []

    for run in runs:
        steps = _extract_ordered_steps(run["graph"])
        exec_data = run.get("execution_data") or {}
        user_map = run.get("user_map") or {}

        for step in steps:
            step_exec = exec_data.get(step["step_id"], {})
            if not isinstance(step_exec, dict):
                step_exec = {}
            results = step_exec.get("results", {})

            base = _build_step_base(run, step, step_exec, user_map)

            original_results = step_exec.get("original_results", {})
            edited = bool(original_results)
            edited_by = user_map.get(
                step_exec.get("edited_by_user_id", ""), ""
            )
            edited_at = step_exec.get("edited_at", "")

            if results:
                for param_name, param_value in results.items():
                    row = dict(base)
                    row["param_name"] = param_name
                    row["param_value"] = (
                        param_value if param_value is not None else ""
                    )
                    row["edited"] = edited
                    row["original_value"] = (
                        original_results.get(param_name, "")
                        if edited else ""
                    )
                    row["edited_by"] = edited_by if edited else ""
                    row["edited_at"] = edited_at if edited else ""
                    rows.append(row)
            elif step_exec.get("value") is not None:
                # Legacy single-value field
                row = dict(base)
                row["param_name"] = "value"
                row["param_value"] = step_exec["value"]
                row["edited"] = edited
                row["original_value"] = (
                    step_exec.get("original_value", "") if edited else ""
                )
                row["edited_by"] = edited_by if edited else ""
                row["edited_at"] = edited_at if edited else ""
                rows.append(row)
            else:
                # Step with no recorded results
                row = dict(base)
                row["param_name"] = ""
                row["param_value"] = ""
                row["edited"] = False
                row["original_value"] = ""
                row["edited_by"] = ""
                row["edited_at"] = ""
                rows.append(row)

    return columns, rows


def _build_wide_format(
    runs: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """One row per step, parameters as dynamic columns."""
    # First pass: collect all unique param names in insertion order
    all_param_names: list[str] = []
    seen_params: set[str] = set()

    for run in runs:
        steps = _extract_ordered_steps(run["graph"])
        exec_data = run.get("execution_data") or {}

        for step in steps:
            step_exec = exec_data.get(step["step_id"], {})
            if not isinstance(step_exec, dict):
                continue
            results = step_exec.get("results", {})
            for param_name in results:
                if param_name not in seen_params:
                    seen_params.add(param_name)
                    all_param_names.append(param_name)
            if not results and step_exec.get("value") is not None:
                if "value" not in seen_params:
                    seen_params.add("value")
                    all_param_names.append("value")

    # Build column list with dynamic param columns in the middle
    param_columns = [
        {"key": f"param__{p}", "label": p, "group": "data"}
        for p in all_param_names
    ]
    columns = WIDE_BASE_COLUMNS + param_columns + WIDE_AUDIT_COLUMNS

    # Second pass: build rows
    rows: list[dict[str, Any]] = []

    for run in runs:
        steps = _extract_ordered_steps(run["graph"])
        exec_data = run.get("execution_data") or {}
        user_map = run.get("user_map") or {}

        for step in steps:
            step_exec = exec_data.get(step["step_id"], {})
            if not isinstance(step_exec, dict):
                step_exec = {}
            results = step_exec.get("results", {})

            base = _build_step_base(run, step, step_exec, user_map)

            original_results = step_exec.get("original_results", {})
            edited = bool(original_results)
            edited_by = user_map.get(
                step_exec.get("edited_by_user_id", ""), ""
            )
            edited_at = step_exec.get("edited_at", "")

            row = dict(base)
            row["edited"] = edited
            row["edited_by"] = edited_by if edited else ""
            row["edited_at"] = edited_at if edited else ""

            # Fill dynamic param columns
            for p in all_param_names:
                key = f"param__{p}"
                if results:
                    val = results.get(p, "")
                    row[key] = val if val is not None else ""
                elif p == "value" and step_exec.get("value") is not None:
                    row[key] = step_exec["value"]
                else:
                    row[key] = ""

            rows.append(row)

    return columns, rows


def filter_columns(
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    selected_keys: list[str] | None,
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """Filter columns and row data to only selected keys."""
    if selected_keys is None:
        return columns, rows

    key_set = set(selected_keys)
    filtered_cols = [c for c in columns if c["key"] in key_set]
    filtered_rows = [
        {k: v for k, v in row.items() if k in key_set}
        for row in rows
    ]
    return filtered_cols, filtered_rows
